import asyncio
import io
import json
import os
import re
import zipfile
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Any

import uvicorn
from fastapi import FastAPI, File, Form, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook as OpenpyxlWorkbook, load_workbook
from pydantic import BaseModel

def now_iso():
    return datetime.now().isoformat()


def _make_empty_xlsx() -> bytes:
    wb = OpenpyxlWorkbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def get_sheet_xml_filename(xlsx_bytes: bytes, sheet_name: str) -> str:
    """
    workbook.xml + workbook.xml.rels 파싱해서
    시트명 → xl/worksheets/sheetN.xml 경로 반환
    """
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")
        sheet_entries = re.findall(
            r'<sheet\s[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"[^>]*/?>',
            workbook_xml
        )
        rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rel_map = dict(re.findall(
            r'Id="([^"]+)"[^>]*Target="([^"]+)"',
            rels_xml
        ))

    for name, rid in sheet_entries:
        if name == sheet_name:
            target = rel_map.get(rid, "")
            if not target.startswith("xl/"):
                target = f"xl/{target}"
            return target

    raise ValueError(
        f"시트명 '{sheet_name}'을 찾을 수 없습니다. "
        f"존재하는 시트: {[n for n, _ in sheet_entries]}"
    )


def build_new_sheet_xml(data: list[dict[str, Any]]) -> bytes:
    """
    JSON 배열 → openpyxl 임시 workbook → sheet XML bytes 추출
    """
    if not data:
        raise ValueError("data 가 비어 있습니다.")

    tmp_wb = load_workbook(io.BytesIO(_make_empty_xlsx()))
    tmp_ws = tmp_wb.active
    headers = list(data[0].keys())

    for col_idx, header in enumerate(headers, start=1):
        tmp_ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            tmp_ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

    buf = io.BytesIO()
    tmp_wb.save(buf)
    buf.seek(0)

    with zipfile.ZipFile(buf) as zf:
        return zf.read("xl/worksheets/sheet1.xml")


def replace_sheet_in_xlsx(
    xlsx_bytes: bytes,
    sheet_name: str,
    new_data: list[dict[str, Any]]
) -> bytes:
    target_path = get_sheet_xml_filename(xlsx_bytes, sheet_name)
    new_sheet_xml = build_new_sheet_xml(new_data)

    input_buf = io.BytesIO(xlsx_bytes)
    output_buf = io.BytesIO()

    with zipfile.ZipFile(input_buf, "r") as zin, \
         zipfile.ZipFile(output_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == target_path:
                zout.writestr(item, new_sheet_xml)
            elif item.filename == "xl/calcChain.xml":
                pass  # ← 추가: calcChain 제거
            elif item.filename == "xl/workbook.xml":
                # ← 추가: 강제 재계산 플래그
                wb_xml = zin.read(item.filename).decode("utf-8")
                if "fullCalcOnLoad" not in wb_xml:
                    if "calcPr" in wb_xml:
                        wb_xml = re.sub(
                            r"<calcPr([^/]*?)/>",
                            r'<calcPr\1 fullCalcOnLoad="1"/>',
                            wb_xml
                        )
                    else:
                        wb_xml = wb_xml.replace(
                            "</workbook>",
                            '<calcPr fullCalcOnLoad="1"/></workbook>'
                        )
                zout.writestr(item, wb_xml.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(item.filename))

    output_buf.seek(0)
    return output_buf.read()


# ────────────────────────────────────────────
# Lifespan
# ────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    app.state.lock = asyncio.Lock()
    yield


app = FastAPI(lifespan=lifespan)


# ────────────────────────────────────────────
# 엔드포인트
# ────────────────────────────────────────────
@app.post("/excel/replace-sheet")
async def replace_excel_sheet(
    file: UploadFile = File(..., description="원본 .xlsx 파일"),
    sheet_name: str = Form(..., description="교체할 시트명 (예: Sheet1)"),
    data: str = Form(..., description="새로 입력할 데이터 JSON 문자열 (배열 형태)"),
):
    try:
        xlsx_bytes = await file.read()

        try:
            parsed_data: list[dict[str, Any]] = json.loads(data)
        except json.JSONDecodeError as e:
            return {"success": False, "error": f"data JSON 파싱 실패: {str(e)}", "timestamp": now_iso()}

        if not isinstance(parsed_data, list) or not parsed_data:
            return {"success": False, "error": "data 는 비어있지 않은 JSON 배열이어야 합니다.", "timestamp": now_iso()}

        result_bytes = await asyncio.get_event_loop().run_in_executor(
            None, replace_sheet_in_xlsx, xlsx_bytes, sheet_name, parsed_data
        )

        original_name = file.filename or "output.xlsx"
        base_name = original_name.rsplit(".", 1)[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return_filename = f"{base_name}_updated_{timestamp}.xlsx"

        print(f"[replace-sheet] 완료 | 시트: '{sheet_name}' | 행수: {len(parsed_data)} | 파일: {return_filename}")

        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{return_filename}"'}
        )

    except ValueError as e:
        return {"success": False, "error": str(e), "timestamp": now_iso()}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"success": False, "error": str(e), "timestamp": now_iso()}


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8027)
